from abc import ABC, abstractmethod
from openpyxl import Workbook, load_workbook


class Product(ABC):
    def __init__(self, item_id, item_name):
        self.item_id = item_id
        self.item_name = item_name

    @abstractmethod
    def calculate_item_amount(self):
        pass


class InventoryItem(Product):
    def __init__(self, item_id, item_name, stock, unit, price, supplier, low_stock_threshold):
        super().__init__(item_id, item_name)
        self.stock = stock
        self.unit = unit
        self.price = price
        self.supplier = supplier
        self.low_stock_threshold = low_stock_threshold

    def calculate_item_amount(self):
        return self.stock * self.price

    def __str__(self):
        return (f"Item ID: {self.item_id}\nItem Name: {self.item_name}\nStock: {self.stock} {self.unit}\n"
                f"Price: {self.price} Tk per {self.unit}\nSupplier: {self.supplier}\n"
                f"Low Stock Threshold: {self.low_stock_threshold}\n")


class Inventory:
    def __init__(self):
        self.inventory = []
        self.new_items_added = False
        self.load_inventory_from_file()

    def load_inventory_from_file(self):
        self.inventory.clear()
        try:
            wb = load_workbook(filename="inventory.xlsx")
            ws = wb.active
            rows = iter(ws.iter_rows(values_only=True))
            next(rows)
            for row in rows:
                item_id, item_name, stock, unit, price, supplier, low_stock_threshold = row
                stock = int(stock)
                price = float(price)
                self.add_item(InventoryItem(item_id, item_name, stock, unit, price, supplier, low_stock_threshold))
            wb.close()
        except FileNotFoundError:
            pass

    def add_item(self, item):
        if not self.__validate_item_id(item.item_id):
            print(f"Invalid Item ID format: {item.item_id}")
            return

        for existing_item in self.inventory:
            if existing_item.item_id == item.item_id:
                print(f"An item with ID {item.item_id} already exists in the inventory. "
                      f"Try adding an item with a unique ID.")
                return

        self.inventory.append(item)
        self.new_items_added = True
        print("Item added to the inventory.")

    def __validate_item_id(self, item_id):
        return isinstance(item_id, int) and item_id > 0

    def save_inventory_to_file(self):
        wb = Workbook()
        ws = wb.active
        headings = ["Item ID", "Item Name", "Stock", "Unit", "Price", "Supplier", "Low Stock Threshold"]
        ws.append(headings)

        for item in self.inventory:
            if isinstance(item, InventoryItem):
                ws.append([item.item_id, item.item_name, item.stock, item.unit, item.price, item.supplier,
                           item.low_stock_threshold])

        wb.save(filename="inventory.xlsx")

    def generate_inventory_summary(self):
        print("Inventory Summary:")
        print("=================")
        for item in self.inventory:
            if isinstance(item, InventoryItem):
                print(item)
                print(f"Total amount for {item.item_name}: {item.calculate_item_amount()} Tk")
                print("-----------------")
        print(f"Total inventory amount: {self.calculate_total_inventory_amount()} Tk")

    def display_inventory(self):
        for item in self.inventory:
            if isinstance(item, InventoryItem):
                print(item)

    def check_low_stock(self):
        low_stock_flag = False

        for item in self.inventory:
            if isinstance(item, InventoryItem) and item.stock < item.low_stock_threshold:
                print(
                    f"Item ID {item.item_id} ({item.item_name}) has low stock. Current stock: {item.stock} {item.unit}")
                low_stock_flag = True

        if not low_stock_flag:
            print("All items have enough stock.")

    def calculate_total_inventory_amount(self):
        total_amount = 0
        for item in self.inventory:
            if isinstance(item, InventoryItem):
                total_amount += item.calculate_item_amount()
        return total_amount

    def delete_item(self, item_id):
        for item in self.inventory:
            if isinstance(item, InventoryItem) and item.item_id == item_id:
                self.inventory.remove(item)
                return True
        return False

    def search_item_by_id(self, item_id):
        for item in self.inventory:
            if isinstance(item, InventoryItem) and item.item_id == item_id:
                return item
        return None

    def search_item_by_name(self, item_name):
        for item in self.inventory:
            if isinstance(item, InventoryItem) and item.item_name.lower() == item_name.lower():
                return item
        return None

    def update_stock(self, item_id, new_stock):
        for item in self.inventory:
            if isinstance(item, InventoryItem) and item.item_id == item_id:
                item.stock = new_stock
                return True
        return False


class User:
    def __init__(self, username, password):
        self.username = username
        self.password = password


def login():
    predefined_user = User("admin", "password")
    print("*********************************************************")
    print("**   InventTrent - Where Inventory Meets Intelligence  **")
    print("*********************************************************")

    username = input("Enter username to login: ")
    password = input("Enter password: ")

    if username == predefined_user.username and password == predefined_user.password:
        print("Login successful!")
        return True
    else:
        print("Invalid username or password. Please try again.")
        return False


def main():
    while not login():
        pass

    inventory = Inventory()

    while True:
        print("*********************************************************")
        print("**   InventTrent - Where Inventory Meets Intelligence  **")
        print("*********************************************************")
        print("**         1. Add Item                                 **")
        print("**         2. Update Stock                             **")
        print("**         3. Delete Item                              **")
        print("**         4. Search Item                              **")
        print("**         5. Display Inventory                        **")
        print("**         6. Check Low Stock Alerts                   **")
        print("**         7. Generate Inventory Summary               **")
        print("**         8. Exit                                     **")
        print("*********************************************************")

        try:
            choice = int(input("Enter your choice: "))
        except ValueError:
            print("Invalid input. Please enter a number between 1 and 8.")
            continue

        if choice == 1:
            try:
                item_id = int(input("Enter Item ID: "))
            except ValueError:
                print("Invalid Item ID. It should be an integer.")
                continue

            item_name = input("Enter Item Name: ")

            try:
                stock_input = input("Enter Stock (with unit, e.g., '50 kg'): ")
                stock, unit = stock_input.split(maxsplit=1)
                stock = int(stock)
            except ValueError:
                print("Invalid stock input. Ensure you enter the stock followed by the unit (e.g., '50 kg').")
                continue

            try:
                price = float(input(f"Enter Price per {unit}: "))
            except ValueError:
                print("Invalid price. It should be a number.")
                continue

            supplier = input("Enter Supplier Information: ")

            try:
                low_stock_threshold = int(input("Enter Low Stock Threshold: "))
            except ValueError:
                print("Invalid low stock threshold. It should be an integer.")
                continue

            new_item = InventoryItem(item_id, item_name, stock, unit, price, supplier, low_stock_threshold)
            inventory.add_item(new_item)

        elif choice == 2:
            try:
                item_id = int(input("Enter Item ID to update stock: "))
                new_stock = int(input("Enter new stock value: "))
            except ValueError:
                print("Invalid input. Item ID and new stock value should be integers.")
                continue

            if inventory.update_stock(item_id, new_stock):
                print("Stock updated.")
            else:
                print("Item not found.")

        elif choice == 3:
            try:
                item_id = int(input("Enter Item ID to delete: "))
            except ValueError:
                print("Invalid Item ID. It should be an integer.")
                continue

            if inventory.delete_item(item_id):
                print("Item deleted from the inventory.")
            else:
                print("Item not found.")

        elif choice == 4:
            try:
                search_choice = int(input("Search by:\n1. Item ID\n2. Item Name\nEnter your choice: "))
            except ValueError:
                print("Invalid choice. Please enter 1 or 2.")
                continue

            if search_choice == 1:
                try:
                    item_id = int(input("Enter Item ID to search: "))
                except ValueError:
                    print("Invalid Item ID. It should be an integer.")
                    continue

                found_item = inventory.search_item_by_id(item_id)
            elif search_choice == 2:
                item_name = input("Enter Item Name to search: ")
                found_item = inventory.search_item_by_name(item_name)
            else:
                print("Invalid choice. Please try again.")
                continue

            if found_item:
                print("Item found:")
                print(found_item)
            else:
                print("Item not found.")

        elif choice == 5:
            inventory.display_inventory()

        elif choice == 6:
            inventory.check_low_stock()

        elif choice == 7:
            inventory.generate_inventory_summary()

        elif choice == 8:
            inventory.save_inventory_to_file()
            break

        else:
            print("Invalid choice. Please try again.")


InventoryManagement = main()
